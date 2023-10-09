from openpyxl import load_workbook

excel_filepath_1 = r"C:\Users\pc\Desktop\AIU\data test\patient_data.xlsx"
excel_filepath_2 = r"C:\Users\pc\Desktop\AIU\data test\encounter.xlsx"

def get_column_data(xl_file_1, xl_file_2) -> tuple:
    """
    Accesses the specified excel files and cleans and stores the data for the specified column.
    
    Args:
        xl_file_1 -> path to the first excel file.
        xl_file_2 -> path to the second excel file.
    """
    sheet_name_file1 ="patient_data (2)"
    sheet_name_file2 = "form_encounter (1)" 

    wb_xl_file_1 = load_workbook(xl_file_1)
    wb_xl_file_2 = load_workbook(xl_file_2)
    active_xl_file_1 = wb_xl_file_1[sheet_name_file1]
    active_xl_file_2 = wb_xl_file_2[sheet_name_file2]

    column_no_1 = 42 # Note: The first row or column integer is 1, not 0.
    column_no_2 = 6 # To get the column number use the excel formula "=COLUMN(A1)" A1 = Select the cell 

    xl_file_1_column_no_1_data = []
    xl_file_2_column_no_2_data = []

    for rows in active_xl_file_1.iter_rows(min_row=3, values_only=True):
        xl_file_1_column_no_1_data.append(rows[column_no_1 - 1]) # For columns with a string datatype clean the data.

    for rows in active_xl_file_2.iter_rows(min_row=3, values_only=True):
        xl_file_2_column_no_2_data.append(rows[column_no_2 - 1])

    return xl_file_1_column_no_1_data, xl_file_2_column_no_2_data

def compare_lists(all_lists) -> None:
    """
    Compares the values of the returned value of the above function for comparison and prints out the values existing
    in one but not the other.

    Args:
        all_lists -> the return type of the get_column_data() function.

    """

    found_values = []
    for values in all_lists[0]: # Compares bothe lists but only returns unique values from one list.
        if values in all_lists[1]:
            found_values.append(values)
    return found_values

    """
    Compares both lists in the tuple and prints only the unque values.

    unique_to_list1 = set(all_lists[0]) - set(all_lists[1])
    unique_to_list2 = set(all_lists[1]) - set(all_lists[0])

    print(unique_to_list1)
    print(unique_to_list2)
    """

def insert_matches(filename, insert_data) -> None:
    """
    The function writes only the values present in file but the other in the active sheet.

    N/B -> Specify the column to write to in the column variable.
    """
    wb_xl_file_2 = load_workbook(filename)
    sheet_name_file2 = "form_encounter (1)" 
    active_xl_file_2 = wb_xl_file_2[sheet_name_file2]
    workbook_to_write = "test.xlsx"

    for index, elements in enumerate(insert_data, start=1):
        active_xl_file_2.cell(row=index, column=32, value=elements)

    wb_xl_file_2.save(workbook_to_write)

"""
# If a number series is needed
def create_No_series(base_value, input_number):
    numeric_part = int(base_value[3:])  # Slices the var and Extracts "0000" and converts it to an integer
    numeric_part += input_number
    formatted_numeric_part = f'{numeric_part:04d}'  # Pads with leading zeros to make it 4 digits
    result = base_value[:3] + formatted_numeric_part
    
    return result
"""





def main():
    matched_data = compare_lists(all_lists= get_column_data(xl_file_1= excel_filepath_1, xl_file_2= excel_filepath_2))
    """insert_matches(
        filename = excel_filepath_2, insert_data= matched_data
    )"""
    print(len(matched_data))
    for int_values in matched_data:
        new_nos = create_No_series(
            base_value= "PAT0000",
            input_number= int_values
        )
        with open('output.txt', 'a') as file:
            file.write(f"{new_nos}\n")


  

if __name__ == "__main__":
    main()